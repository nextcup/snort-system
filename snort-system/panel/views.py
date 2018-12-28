#!/usr/bin/python
# -*- coding: UTF-8 -*-


from django.shortcuts import render_to_response
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.views.generic import TemplateView, ListView
from django.http import StreamingHttpResponse
from enum import Enum, unique
from openpyxl import Workbook
from update import *

import hashlib
import zipfile
import random
import commands
import datetime
import json


# Create your views here.


class Show(ListView):
    """
    规则概要展示
    """
    template_name = 'rules.html'
    paginate_by = 10

    def get_queryset(self):
        """
        :param:   类对象
        :return:  数据库特征表实例
        """
        rules_summary = Rule.objects.all()
        return rules_summary

    def set_data(self):
        """
        :param:   类对象
        :return:  分页对象, 每页数据
        """
        rules_summary = self.get_queryset()
        paginator = Paginator(rules_summary, 10)
        page = self.request.GET.get('page', 1)
        result = paginator.page(page)
        return paginator, result

    def get_context_data(self, **kwargs):
        """
        :param param1: 类对象
        :param param2: 可变参
        :return:       将分页对象和每页数据通过context封装返回
        """
        context = super(Show, self).get_context_data(**kwargs)
        context['paginator'], context['rules_sections'] = self.set_data()
        return context


class LogShow(ListView):
    """
    日志展示 
    """
    template_name = 'log.html'
    paginate_by = 10

    def get_queryset(self):
        """
        :param:  类对象
        :return: 数据库日志表实例
        """
        logs = Log.objects.all()
        return logs

    def set_data(self):
        """
        :param:  类对象
        :return: 分页对象, 每页数据
        """
        logs = self.get_queryset()
        paginator = Paginator(logs, 10)
        page = self.request.GET.get('page', 1)
        result = paginator.page(page)
        return paginator, result

    def get_context_data(self, **kwargs):
        """
        :param param1:  类对象
        :param param2:  可变参
        :return:        将分页对象和每页数据通过context封装返回
        """
        context = super(LogShow, self).get_context_data(**kwargs)
        context['paginator'], context['rules_sections'] = self.set_data()
        return context


class Untranslate(ListView):
    """
    未翻译界面
    """
    template_name = 'untranslate.html'
    paginate_by = 10

    def get_queryset(self):
        """
        :param:  类对象
        :return: 数据库中未翻译的规则集合实例  
        """
        rules_summary = Rule.objects.filter(is_translate='否')
        return rules_summary

    def set_data(self):
        """
        :param:  类对象
        :return: 分页对象, 每页数据
        """
        rules_summary = self.get_queryset()
        paginator = Paginator(rules_summary, 10)
        page = self.request.GET.get('page', 1)
        result = paginator.page(page)
        return paginator, result

    def get_context_data(self, **kwargs):
        """
        :param param1:  类对象
        :param param2:  可变参
        :return:        将分页对象和每页数据通过context封装返回
        """
        context = super(Untranslate, self).get_context_data(**kwargs)
        context['paginator'], context['rules_sections'] = self.set_data()
        return context


class ConflictShow(ListView):
    """
    冲突列表展示 
    """
    template_name = 'conflict.html'
    paginate_by = 10

    def get_queryset(self):
        """
        :param:  类对象
        :return: 数据库中冲突的规则集合实例  
        """
        rules = Rule.objects.filter(has_conflict='是')
        return rules

    def set_data(self):
        """
        :param:  类对象
        :return: 分页对象, 每页数据
        """
        rules = self.get_queryset()
        paginator = Paginator(rules, 10)
        page = self.request.GET.get('page', 1)
        result = paginator.page(page)
        return paginator, result

    def get_context_data(self, **kwargs):
        """
        :param param1:  类对象
        :param param2:  可变参
        :return:        将分页对象和每页数据通过context封装返回
        """
        context = super(ConflictShow, self).get_context_data(**kwargs)
        context['paginator'], context['rules_sections'] = self.set_data()
        return context


class SearchResult(object):
    """
    规则特征查询
    """

    def __init__(self):
        self.rules_summary = None
        self.search_data = get_search_data()
        self.start, self.end = get_start_end(get_date_range())
        self.option = get_option()

    def deal_not_chinese(self):
        """
        :param:  类对象
        :return: 用户输入的查询数据
        """
        if not has_chinese(self.search_data):
            self.search_data = str(self.search_data).strip()
            return self.search_data
        return self.search_data

    def option_is_sid(self):
        """
        :param:  类对象
        :return: 规则ID查询结果
        """
        self.search_data = self.deal_not_chinese()
        if self.start == "" and self.end == "":
            self.rules_summary = Rule.objects.filter(
                sid__contains=self.search_data)
        else:
            self.rules_summary = Rule.objects.filter(
                sid__contains=self.search_data, update_time__gte=self.start, update_time__lte=self.end)
        return self.rules_summary

    def option_is_msg(self):
        """
        :param:  类对象
        :return: 规则描述信息查询结果
        """
        self.search_data = self.deal_not_chinese()
        if self.start == "" and self.end == "":
            self.rules_summary = Rule.objects.filter(
                msg__contains=self.search_data)
        else:
            self.rules_summary = Rule.objects.filter(
                msg__contains=self.search_data, update_time__gte=self.start, update_time__lte=self.end)
        return self.rules_summary

    def option_is_classtype(self):
        """
        :param:  类对象
        :return: 规则病毒类型查询结果
        """
        self.search_data = self.deal_not_chinese()
        if self.start == "" and self.end == "":
            self.rules_summary = Rule.objects.filter(
                class_type__contains=self.search_data)
        else:
            self.rules_summary = Rule.objects.filter(
                class_type__contains=self.search_data, update_time__gte=self.start, update_time__lte=self.end)
        return self.rules_summary

    def option_is_malname(self):
        """
        :param:  类对象
        :return: 规则恶意家族查询结果
        """
        self.search_data = self.deal_not_chinese()
        if self.start == "" and self.end == "":
            self.rules_summary = Rule.objects.filter(
                malname__contains=self.search_data)
        else:
            self.rules_summary = Rule.objects.filter(
                malname__contains=self.search_data, update_time__gte=self.start, update_time__lte=self.end)
        return self.rules_summary

    def option_is_date_range(self):
        """
        :param:  类对象
        :return: 规则修改时间范围查询结果
        """
        self.rules_summary = Rule.objects.filter(
            update_time__gte=self.start, update_time__lte=self.end)
        return self.rules_summary


class UntransSearchResult(object):
    """
    未翻译规则查询
    """

    def __init__(self):
        self.rules_summary = None
        self.search_data = get_untrans_search_data()
        self.option = get_untrans_option()

    def deal_not_chinese(self):
        """
        :param:  类对象
        :return: 用户输入的查询数据
        """
        if not has_chinese(self.search_data):
            self.search_data = str(self.search_data).strip()
            return self.search_data
        return self.search_data

    def option_is_sid(self):
        """
        :param:  类对象
        :return: 规则ID查询结果
        """
        self.search_data = self.deal_not_chinese()
        self.rules_summary = Rule.objects.filter(
            sid__contains=self.search_data, is_translate='否')
        return self.rules_summary

    def option_is_msg(self):
        """
        :param:  类对象
        :return: 规则描述信息查询结果
        """
        self.search_data = self.deal_not_chinese()
        self.rules_summary = Rule.objects.filter(
            msg__contains=self.search_data, is_translate='否')
        return self.rules_summary

    def option_is_classtype(self):
        """
        :param:  类对象
        :return: 规则病毒类型查询结果
        """
        self.search_data = self.deal_not_chinese()
        self.rules_summary = Rule.objects.filter(
            class_type__contains=self.search_data, is_translate='否')
        return self.rules_summary

    def option_is_malname(self):
        """
        :param:  类对象
        :return: 规则恶意家族查询结果
        """
        self.search_data = self.deal_not_chinese()
        self.rules_summary = Rule.objects.filter(
            malname__contains=self.search_data, is_translate='否')
        return self.rules_summary


@unique
class Features(Enum):
    """
    特征枚举
    """
    sid = 0                                 # 规则ID
    msg = 1                                 # 描述信息
    reference = 2                           # 引用信息
    class_type = 3                          # 类型
    malname = 4                             # 家族名
    attacker = 5                            # 攻击者
    victim = 6                              # 受害者
    success_attack = 7                      # 成功攻击
    controller = 8                          # 控制源
    confirm_controlled = 9                  # 确认被控
    rev = 10                                # 修订版本
    knowledge_base = 11                     # 知识库信息
    shield = 12                             # 是否屏蔽
    contain = 13                            # 出库包含
    first_check_position = 14               # 首次检出位置
    overall_first_check_position = 15       # 整体首次检出位置
    check_out_numbers = 16                  # 检出个数
    error_numbers = 17                      # 误报个数
    feature_source = 18                     # 特征源
    remarks = 19                            # 备注
    content = 20                            # 匹配信息


class Add(TemplateView):
    """
    新增规则界面
    """
    template_name = 'new_rules.html'


class Shield(TemplateView):
    """
    规则屏蔽界面
    """
    template_name = 'shield.html'


class Edit(TemplateView):
    """
    规则修改界面
    """
    template_name = 'edit.html'


class Upload(TemplateView):
    """
    文件上传界面
    """
    template_name = 'upload.html'


class Download(TemplateView):
    """
    导出规则文件界面
    """
    template_name = 'download.html'


class About(TemplateView):
    """
    系统功能简介界面
    """
    template_name = 'about.html'


class DownloadPcap(TemplateView):
    """
    下载pcap界面
    """
    template_name = 'download_pcap.html'


class DeletePcap(TemplateView):
    """
    删除pcap界面
    """
    template_name = 'delete_pcap.html'


class Xlsx(TemplateView):
    """
    导出xlsx表格界面
    """
    template_name = 'xlsx.html'


def backstage(request):
    """
    :describe:  系统后台界面显示
    :param      http请求
    :return:    系统后台界面
    """
    start_update()
    try:
        user_now = request.session['user_name']
    except Exception as e:
        print 'login error:', e
        return HttpResponse('错误: 请访问http://192.168.18.182:8000/登陆!')
    return render_to_response('backstage.html', {'user': user_now})


@csrf_exempt
def register(request):
    """
    :describe:  用户注册
    :param      http请求
    :return:    登陆、注册界面
    """
    if request.method == 'GET':
        return render_to_response('index.html')

    username = request.POST.get('username_register')
    clear_passwd = request.POST.get('password_register')
    password = encryption(clear_passwd)

    if User.objects.filter(username=username):
        return HttpResponse(0)

    user = User.objects.create(username=username, password=password)
    ip = get_ip(request)
    if user:
        request.session['is_register'] = True
        record_log('无', '用户注册', username, '成功', ip, '用户成功注册')
        return HttpResponse(1)
    else:
        return HttpResponse(0)

    return render_to_response('index.html')


@csrf_exempt
def login(request):
    """
    :describe:  用户登录
    :param      http请求
    :return:    登陆响应值 0:成功 1:失败
    """
    if request.method == 'GET':
        return render_to_response('index.html')

    username = request.POST.get('username')
    password = request.POST.get('password')

    user = User.objects.filter(
        username=username, password=encryption(password))
    ip = get_ip(request)
    if request.session.get('is_login', None):
        if request.session['user_name'] == username:
            return HttpResponse(2)
    if user:
        request.session['is_login'] = True
        request.session['user_name'] = username
        record_log('无', '用户登录', username, '成功', ip, '用户成功登陆')
        return HttpResponse(1)
    else:
        return HttpResponse(0)


def logout(request):
    """
    :describe:  注销
    :param      http请求
    :return:    注销后返回注册、登陆界面
    """
    if not request.session.get('is_login', None):
        return render_to_response('index.html')
    ip = get_ip(request)
    username = request.session['user_name']
    record_log('无', '用户注销', username, '成功', ip, '用户成功注销')
    request.session.flush()

    return render_to_response('index.html')


@csrf_exempt
def detail_show(request):
    """
    :describe:  查看规则详情
    :param      http请求
    :return:    特征详情界面及其包含的数据
    """
    if request.method == 'POST':
        set_sid(request.POST.get('sid'))
        set_rule_data(Rule.objects.get(sid=get_sid()))

    return render_to_response('view_detail.html', {'rule': get_rule_data()})


@csrf_exempt
def complete_show(request):
    """
    :describe:  查看完整规则
    :param      http请求
    :return:    完整规则界面及其包含的数据
    """
    if request.method == 'POST':
        set_sid(request.POST.get('sid'))
        set_rule_data(CompleteRule.objects.get(sid=get_sid()))

    return render_to_response('complete.html', {'rule': get_rule_data()})


def log_detail(request):
    """
    :describe:  查看日志详情
    :param      http请求
    :return:    日志详情界面及其历史日志数据
    """
    index = 1
    msg_str = ""
    final_str = ""

    if request.method == 'POST':
        set_default_env()
        set_sid(request.POST.get('sid'))

        rules_summary = Log.objects.filter(sid=str(get_sid()))

        for v in rules_summary:
            msg_str = str(index) + '.  ' + msg_str + v.msg + \
                '    ' + v.time + '    ' + v.person + '\n\n'
            final_str += msg_str
            msg_str = ""
            index += 1
        set_history_log(final_str)

    return render_to_response('log_detail.html', {'rules_sections': get_history_log()})


@csrf_exempt
def search(request):
    """
    :describe:  特征概要模块查询
    :param      http请求
    :return:    响应值 0:成功 1:失败 成功后跳转结果展示
    """
    if request.method == 'GET':
        return render_to_response('rules.html')

    set_search_data(request.POST.get('search_data'))
    set_option(request.POST.get('option'))
    set_date_range(request.POST.get('date_range'))

    if get_search_data() or get_date_range():
        return HttpResponse(1)

    return HttpResponse(0)


''' 未翻译查询功能 '''


@csrf_exempt
def untrans_search(request):
    """
    :describe:  未翻译模块查询
    :param      http请求
    :return:    响应值 0:成功 1:失败 成功后跳转结果展示
    """
    if request.method == 'GET':
        return render_to_response('untranslate.html')

    set_untrans_search_data(request.POST.get('search_data'))
    set_untrans_option(request.POST.get('option'))

    if get_untrans_search_data():
        return HttpResponse(1)
    return HttpResponse(0)


''' 获取时间范围: start end '''


def get_start_end(date_str):
    """
    :describe:  获取时间范围: start end
    :param      前端传来的时间字符串 2018-12-01 00:00:00 - 2019-01-31 00:00:00
    :return:    start: 2018-12-01 00:00:00 end: 2019-01-31 00:00:00
    """
    if date_str == "":
        return "", ""

    start = datetime.datetime.strptime(
        date_str.split(' - ')[0].strip(), "%Y-%m-%d %H:%M:%S")

    end = datetime.datetime.strptime(
        date_str.split(' - ')[1].strip(), "%Y-%m-%d %H:%M:%S")

    return start, end


def set_page_data(request, rules_summary):
    """
    :param param1:  http请求
    :param param2:  查询结果
    :return:        分页对象及数据
    """
    paginator = Paginator(rules_summary, 10)
    page = request.GET.get('page', 1)
    result = paginator.page(page)
    return paginator, result


def search_result(request):
    """
    :describe:  特征概要查询结果展示
    :param      http请求
    :return:    查询结果且分页展示
    """
    rules_summary = None
    search_obj = SearchResult()
    set_default_env()

    if 'sid' in search_obj.option:
        rules_summary = search_obj.option_is_sid()
    elif 'msg' in search_obj.option:
        rules_summary = search_obj.option_is_msg()
    elif 'class_type' in search_obj.option:
        rules_summary = search_obj.option_is_classtype()
    elif 'malname' in search_obj.option:
        rules_summary = search_obj.option_is_malname()
    elif '时间范围' in search_obj.option:
        rules_summary = search_obj.option_is_date_range()

    if rules_summary is None:
        return render_to_response('result.html', {'rules_sections': None})

    paginator, result = set_page_data(request, rules_summary)

    return render_to_response('result.html', {'rules_sections': result, 'paginator': paginator})


def untrans_result(request):
    """
    :describe:  未翻译模块查询结果展示
    :param      http请求
    :return:    查询结果且分页展示
    """
    rules_summary = None
    untrans_search_obj = UntransSearchResult()

    if 'sid' in untrans_search_obj.option:
        rules_summary = untrans_search_obj.option_is_sid()
    elif 'msg' in untrans_search_obj.option:
        rules_summary = untrans_search_obj.option_is_msg()
    elif 'class_type' in untrans_search_obj.option:
        rules_summary = untrans_search_obj.option_is_classtype()
    elif 'malname' in untrans_search_obj.option:
        rules_summary = untrans_search_obj.option_is_malname()

    paginator, result = set_page_data(request, rules_summary)

    return render_to_response('untrans_result.html', {'rules_sections': result, 'paginator': paginator})


@csrf_exempt
def add_content(request):
    """
    :describe:  添加多条content字段请求
    :param      http请求
    :return:    响应值 0:失败 1:成功
    """
    if request.method == "POST":
        try:
            receive_data = request.POST.get('another_content')
            status = request.POST.get('status')
            receive_data = json.loads(receive_data)
            set_another_content_list(receive_data)
            set_status(status)
        except Exception as e:
            print 'error: ', e
            return HttpResponse(0)

    return HttpResponse(1)


def get_receive_data(request):
    receive_data = None
    user = request.session["user_name"]
    if request.method == "POST":
        receive_data = request.POST.get('data')
        receive_data = json.loads(receive_data)
    receive_data = combina_content(receive_data)
    if get_status() == "success":
        receive_data.pop("content")
    return receive_data, user


def get_final_rule(request, sid):
    sub_rule = ""
    final_rule = ""
    receive_data = None
    no_loop_key = ['protocol', 'src_ip', 'src_port', 'dst_ip', 'dst_port']
    receive_data, user = get_receive_data(request)
    try:
        rule_str = 'alert ' + receive_data["protocol"] + ' ' + receive_data["src_ip"] + ' ' + receive_data["src_port"] + ' -> ' + \
            receive_data["dst_ip"] + ' ' + receive_data["dst_port"] + ' ('
    except Exception as e:
        print 'error:', e
        return HttpResponse(0)
    for k in no_loop_key:
        receive_data.pop(k)
    for key, val in receive_data.items():
        if key == "msg" or key == "pcre":
            sub_rule = sub_rule + key + ':"' + val + '";'
        elif key == "content":
            sub_rule = sub_rule + val
        else:
            sub_rule = sub_rule + key + ':' + val + ';'
    final_rule = rule_str + sub_rule + 'sid:' + str(sid) + ';'
    another_content = get_another_content_list()
    if another_content is not None:
        for val in another_content:
            final_rule += val["content"]
        final_rule += ')'
    return final_rule, receive_data, user


@csrf_exempt
def add_submit(request):
    """
    :describe:  新增规则并检验规则
    :param      http请求
    :return:    响应值 0:失败 1:成功 2:规则不支持
    """
    sid = random_sid()
    ip = get_ip(request)
    final_rule, receive_data, user = get_final_rule(request, sid)
    # 规则验证
    new_rule_path = get_pcap_path() + 'new.rules'
    with open(new_rule_path, 'w')as f:
        f.write(final_rule + '\n')
    if len(get_another_content_list()) == 0:
        return HttpResponse(3)
    if not rules_verify(new_rule_path):
        os.system('rm -rf ' + new_rule_path)
        return HttpResponse(2)
    os.system('rm -rf ' + new_rule_path)
    # 存储特征
    if not syn_features(sid, receive_data, user, ip):
        return HttpResponse(0)
    # 存储规则
    if not syn_new_rule(sid, final_rule, user, ip):
        return HttpResponse(0)
    reset_content()
    return HttpResponse(1)


def get_add_content():
    """
    :describe:  获取新增content字段
    :param      无
    :return:    content拼接结果
    """
    new_content_str = ""
    for val in get_another_content_list():
        new_content_str += val["content"] + '\n'
    return new_content_str


def combina_content(data_dict):
    """
    :describe:  content及其选项组合
    :param      前端返回字典数据
    :return:    字典中content字段和选项字段组合为一个字段
    """
    options = ['offset', 'depth', 'nocase', 'distance', 'within']
    option_str = ""
    start_content = ""

    for key, val in data_dict.items():
        if key == 'content':
            start_content = key + ':"' + val + '";'
        elif key in options and val != "":
            if key == "nocase" and val == "true":
                option_str += key + ';'
            elif key == "nocase" and val == "false":
                pass
            else:
                option_str += key + ':' + val + ';'
        else:
            continue
    data_dict["content"] = start_content + option_str
    for val in options:
        if val in data_dict.keys():
            data_dict.pop(val)
        else:
            continue
    return data_dict


def syn_new_rule(sid, new_rule, user, ip):
    """
    :describe:      规则添加后同步规则入库
    :param param1:  规则ID
    :param param2:  规则字符串
    :param param3:  当前登录用户
    :return:        True:同步成功 False:同步失败
    """
    try:
        rule_obj = CompleteRule.objects.create(sid=str(sid), rule=new_rule)
        rule_obj.save()
        record_log(str(sid), '新增规则', user, '成功', ip, '成功新增一条snort规则并且规则成功入库')
        return True
    except Exception as e:
        record_log(str(sid), '新增规则', user, '失败', ip, '错误信息:%s' % (e))
        return False


def syn_features(sid, receive_data, user, ip):
    """
    :describe:      规则添加后同步特征入库
    :param param1:  规则ID
    :param param2:  前端返回字典数据(key:特征,value:具体值)
    :param param3:  当前登录用户
    :return:        True:同步成功 False:同步失败
    """
    if 'reference' in receive_data.keys():
        refs = receive_data['reference']
    else:
        refs = ""
    try:
        rule_obj = Rule.objects.create(
            sid=str(sid),
            msg=receive_data["msg"],
            reference=refs,
            class_type=receive_data["classtype"],
            malname=receive_data["classtype"] + '/Generic.' + str(sid),
            attacker='',
            victim='',
            success_attack='',
            controller='',
            confirm_controlled='',
            rev=receive_data["rev"],
            create_time=get_date(),
            update_time=get_date(),
            is_translate='否',
            content=get_add_content(),
            has_conflict='否'
        )
        rule_obj.save()
        record_log(str(sid), '新增规则后特征入库', user, '成功',
                   ip, '成功新增一条snort规则并且特征成功入库')
        return True
    except Exception as e:
        record_log(str(sid), '新增规则后特征入库', user, '失败', ip, '错误信息:%s' % (e))
        return False


@csrf_exempt
def shield(request):
    """
    :describe: 规则屏蔽
    :param:    http请求 
    :return:   响应值: 0:失败 1:成功
    """
    # 屏蔽指定规则
    sid = request.POST.get('id')
    rule_obj = Rule.objects.get(id=sid)
    user = request.session['user_name']
    ip = get_ip(request)

    complete_rule = None
    try:
        complete_rule = CompleteRule.objects.get(sid=rule_obj.sid)
    except Exception as e:
        Rule.objects.filter(sid=rule_obj.sid).update(shield='否')
        record_log(rule_obj.sid, '规则屏蔽', user, '失败', ip, e)
        return HttpResponse(0)

    if complete_rule.rule.startswith('#'):
        record_log(rule_obj.sid, '规则屏蔽', user, '失败', ip, '该规则重复屏蔽')
        return HttpResponse(0)

    CompleteRule.objects.filter(sid=rule_obj.sid).update(
        rule="#" + complete_rule.rule)
    Rule.objects.filter(sid=rule_obj.sid).update(shield='是')
    record_log(rule_obj.sid, '规则屏蔽', user, '成功', ip, '规则屏蔽成功')

    return HttpResponse(1)


def get_ip(request):
    """
    :describe: 获取访问ip
    :param:    http请求 
    :return:   ip地址
    """
    ip = None
    if request.META.has_key('HTTP_X_FORWARDED_FOR'):
        ip = request.META['HTTP_X_FORWARDED_FOR']
    else:
        ip = request.META['REMOTE_ADDR']
    return ip


@csrf_exempt
def edit(request):
    """
    :describe: 规则特征修改
    :param:    http请求 
    :return:   响应值: 0:失败 1:成功
    """
    # 弹窗显示规则指定字段,可编辑并保存
    if request.method == 'GET':
        return render_to_response('rules.html')

    data = request.POST.getlist('data')
    user = request.session['user_name']
    ip = get_ip(request)
    flag = request.POST.get('flag')
    old_features_dict = {}

    if data:
        # 记录修改之前的特征
        old_features_dict = Rule.objects.filter(sid=str(data[0])).values()[0]
        # 将修改后内容插入数据库
        Rule.objects.filter(sid=data[Features.sid.value]).update(
            msg=data[Features.msg.value],
            reference=data[Features.reference.value],
            class_type=data[Features.class_type.value],
            malname=data[Features.malname.value],
            attacker=data[Features.attacker.value],
            victim=data[Features.victim.value],
            success_attack=data[Features.success_attack.value],
            controller=data[Features.controller.value],
            confirm_controlled=data[Features.confirm_controlled.value],
            rev=str(int(data[Features.rev.value]) + 1),
            knowledge_base=data[Features.knowledge_base.value],
            shield=data[Features.shield.value],
            contain=data[Features.contain.value],
            first_check_position=data[Features.first_check_position.value],
            overall_first_check_position=data[Features.overall_first_check_position.value],
            check_out_numbers=data[Features.check_out_numbers.value],
            error_numbers=data[Features.error_numbers.value],
            feature_source=data[Features.feature_source.value],
            remarks=data[Features.remarks.value],
            content=data[Features.content.value],
            update_time=get_date())

        if flag == 'conflict_flag':
            record_log(data[0], '冲突解决', user, '成功', ip, '自动更新发生冲突解决')
            Rule.objects.filter(sid=data[0]).update(has_conflict='否')
            if len(get_edited_rule_id()) != 0:
                get_edited_rule_id().remove(data[0])
            UpdateRule.objects.get(sid=data[0]).delete()
        else:
            pass
        compare(old_features_dict, data, user, ip)
        synchro(data, user, ip)
        return HttpResponse(1)
    return HttpResponse(0)


def compare(old_features, new_features, user, ip):
    """
    :describe:      记录被修改的特征字段
    :param param1:  规则修改之前特征
    :param param2:  规则修改之后特征
    :param param3:  当前登录用户
    :return:        无
    """
    set_default_env()
    for k, v in old_features.items():
        for enum_v in Features:
            if k in str(enum_v):
                if k == 'id' or k == 'rev':
                    continue
                if v != new_features[enum_v.value]:
                    if v != "":
                        record_log(new_features[0], '规则修改', user, '成功', ip, '%s字段由%s修改为%s' % (
                            k, v, new_features[enum_v.value]))
                    else:
                        record_log(new_features[0], '规则修改', user, '成功', ip, '%s字段被设置为%s' % (
                            k, new_features[enum_v.value]))
                else:
                    continue
            else:
                continue


def write_check_rule(sid, rule):
    """
    :describe:      写入要被检测的规则到文件(ID命名)
    :param param1:  规则ID
    :param param2:  规则字符串
    :return:        成功: 文件路径 失败: None
    """
    pcaps_dir = get_upload_path()
    if not os.path.exists(pcaps_dir):
        os.makedirs(pcaps_dir)

    path = pcaps_dir + str(sid) + '.rules'

    try:
        with open(path, 'w') as f:
            f.write(rule)
            return path
    except Exception as e:
        print e
        return None


def stor_rule_pcap(sid, pcap):
    """
    :describe:      存储成功匹配的规则和pcap
    :param param1:  规则ID
    :param param2:  pcap文件
    :return:        无
    """
    rule_obj = RulePcap.objects.create(sid=sid, pcap=pcap)
    rule_obj.save()


@csrf_exempt
def upload(request):
    """
    :describe: pcap上传
    :param:    http请求
    :return:   响应状态: 成功 失败 格式 是否选择 支持情况
    """
    set_default_env()

    if request.method == "POST":
        sid = request.POST.get('sid')
        if sid is not None:
            set_id(sid)
        rule = CompleteRule.objects.get(sid=get_id())
        # 将规则写入文件
        checked_rule = write_check_rule(get_id(), rule.rule)
        if checked_rule is None:
            return HttpResponse('待检测规则写入文件失败!')
        # 调用工具判断规则是否支持
        if not rules_verify(checked_rule):
            return HttpResponse('规则不支持,请修改或屏蔽!')

        result = upload_pcap(request, checked_rule)
        return HttpResponse(result)
    return HttpResponse("上传失败!")


def upload_pcap(request, checked_rule):
    return_str = ""
    myFile = request.FILES.get("myfile", None)
    if not myFile:
        return_str = "未选择文件!"
        return return_str
    if not str(myFile).endswith('.pcap'):
        return_str = "文件格式错误"
        return return_str
    pcap_path = get_upload_path() + str(myFile)

    path = get_upload_path()
    destination = open(os.path.join(path, myFile.name), 'wb+')
    for chunk in myFile.chunks():
        destination.write(chunk)
    destination.close()
    user = request.session['user_name']
    ip = get_ip(request)
    if is_hit(checked_rule, pcap_path, user, ip) is False:
        set_default_env()
        tips = '%s号规则与pcap包%s未命中' % (get_id(), myFile)
        # 删除未命中pcap
        cmd = 'rm -rf ./data/pcaps/' + str(myFile)
        os.system(cmd)
        return tips
    record_log('无', '文件上传', user, '成功', ip,
               '文件为%s,服务器存储路径为%s' % (myFile.name, path))
    # 存储成功匹配的规则和pcap信息
    stor_rule_pcap(get_id(), myFile)
    tips = '%s号规则与pcap包%s成功命中' % (get_id(), myFile)
    return tips


@csrf_exempt
def trans_submit(request):
    """
    :describe:  规则翻译
    :param:     http请求
    :return:    响应值 0:失败 1:成功
    """
    if request.method == 'GET':
        return render_to_response('untranslate.html')

    trans_msg = request.POST.get('msg')
    rule_sid = get_trans_sid()
    rule_obj = Rule.objects.get(sid=rule_sid)
    Rule.objects.filter(sid=rule_sid).update(msg=trans_msg, is_translate='是')
    user = request.session['user_name']
    ip = get_ip(request)
    set_default_env()
    record_log(rule_sid, '规则翻译', user, '成功', ip,
               '将%s翻译为%s' % (rule_obj.msg, trans_msg))

    return HttpResponse(1)


def generate_zip():
    """
    :describe:  压缩要下载的文件
    :param:     无
    :return:    无
    """
    file_list = []
    path = get_download_path()
    for root, dirs, files in os.walk(path, topdown=False):
        for name in files:
            if not name.endswith('.py') \
                    and not name.endswith('.zip') \
                    and not name.endswith('.xlsx'):
                file_list.append(os.path.join(root, name))

    f = zipfile.ZipFile(
        path + 'rules.zip', 'w', zipfile.ZIP_DEFLATED)
    for file in file_list:
        f.write(file)
    f.close()


@csrf_exempt
def download(request):
    """
    :describe:  规则导出
    :param:     http请求
    :return:    导出的文件对象
    """
    db_to_file()
    extract_contain_feature()
    generate_zip()
    path = get_zip_path()

    file = None
    try:
        file = open(path, 'rb')
    except Exception as e:
        print e
        return HttpResponse('待导出文件打开出错')
    response = StreamingHttpResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="rules.zip"'
    user = request.session['user_name']
    ip = get_ip(request)
    set_default_env()
    record_log('无', '规则导出', user, '成功', ip,
               '导出文件为rules.zip,下载路径为%s' % (path))

    return response


def get_new_content(key):
    """
    :describe:  提取更新来的规则content字段
    :param:     规则ID
    :return:    拼接后的content字符串
    """
    content = []
    content_str = ""
    options = ['offset', 'depth', 'nocase', 'distance', 'within']
    rule = UpdateRule.objects.get(sid=int(key)).rule
    result = rule.split(';')
    for val in result:
        if 'content' in val:
            content.append(val)
        else:
            [content.append(val) for v in options if v in val]
    if len(content) == 0:
        return ""
    flag = False
    for val in content:
        val = val.strip()
        if 'content' in val:
            if flag is True:
                content_str += '\n'
                content_str += val + ';'
            else:
                content_str += val + ';'
                flag = True
        else:
            content_str += val + ';'
            flag = True
    return content_str


def deal(request):
    """
    :describe:  冲突数据获取
    :param:     http请求
    :return:    无
    """
    if request.method == 'POST':
        set_sid(request.POST.get('sid'))
        set_rule_data(Rule.objects.get(sid=get_sid()))

        rule_obj = UpdateRule.objects.get(sid=get_sid())

        set_update_sid(re.findall(r'sid:(.+?);', rule_obj.rule))
        set_update_msg(re.findall(r'msg:\"(.+?)\";', rule_obj.rule))
        set_update_reference(re.findall(
            r'reference:(.+?);', rule_obj.rule))
        set_update_content(get_new_content(get_sid()))
        set_update_classtype(re.findall(r'classtype:(.+?);', rule_obj.rule))

    return render_to_response(
        'deal.html', {
            'rule': get_rule_data(),
            'sid': get_update_sid()[0],
            'msg': get_update_msg()[0],
            'reference': get_update_reference(),
            'content': get_update_content(),
            'class_type': get_update_classtype()[0]

        })


@csrf_exempt
def deal_conflict(request):
    """
    :describe:  冲突解决
    :param:     http请求
    :return:    响应值 0:失败 1:成功
    """
    set_default_env()
    if request.method == 'GET':
        return render_to_response('conflict.html')

    data = request.POST.getlist('data')
    user = request.session['user_name']
    ip = get_ip(request)

    if data:
        # 将修改后内容插入数据库
        Rule.objects.filter(sid=data[Features.sid.value]).update(
            msg=data[Features.msg.value],
            reference=data[Features.reference.value],
            class_type=data[Features.class_type.value],
            malname=data[Features.malname.value],
            attacker=data[Features.attacker.value],
            victim=data[Features.victim.value],
            success_attack=data[Features.success_attack.value],
            controller=data[Features.controller.value],
            confirm_controlled=data[Features.confirm_controlled.value],
            rev=data[Features.rev.value],
            knowledge_base=data[Features.knowledge_base.value],
            shield=data[Features.shield.value],
            contain=data[Features.contain.value],
            first_check_position=data[Features.first_check_position.value],
            overall_first_check_position=data[Features.overall_first_check_position.value],
            check_out_numbers=data[Features.check_out_numbers.value],
            error_numbers=data[Features.error_numbers.value],
            feature_source=data[Features.feature_source.value],
            remarks=data[Features.remarks.value],
            content=data[Features.content.value],
            update_time=get_date())

        record_log(data[0], '冲突解决', user, '成功', ip, '自动更新发生冲突解决')
        synchro(data, user)
        Rule.objects.filter(sid=data[0]).update(has_conflict='否')
        if len(get_edited_rule_id()) != 0:
            get_edited_rule_id().remove(data[0])
        UpdateRule.objects.get(sid=data[0]).delete()

        return HttpResponse(1)
    record_log('无', '冲突解决', user, '失败', ip, '数据有误')
    return HttpResponse(0)


@csrf_exempt
def download_pcap(request):
    """
    :describe:  pcap下载
    :param:     http请求
    :return:    下载的文件
    """
    pcap = ""
    sid = request.POST.get('sid')
    if sid is not None:
        set_rule_pcap(sid)

    rules = RulePcap.objects.all()
    if len(rules) == 0:
        return HttpResponse('该规则无命中的pcap')
    for rule in rules:
        if rule.sid == get_rule_pcap():
            pcap = rule.pcap

    file_list = []
    path = get_pcap_path()
    for root, dirs, files in os.walk(path, topdown=False):
        for name in files:
            if name.endswith('.pcap') and pcap in name and pcap != "":
                file_list.append(os.path.join(root, name))
            else:
                continue

    if len(file_list) == 0:
        return HttpResponse('该规则无命中的pcap')

    file = open(file_list[0], 'rb')
    response = StreamingHttpResponse(file)
    response['Content-Type'] = 'application/octet-stream'

    response['Content-Disposition'] = 'attachment;filename="rule.pcap"'
    user = request.session['user_name']
    ip = get_ip(request)
    set_default_env()
    record_log('无', 'pcap下载', user, '成功', ip,
               '导出文件为rule.pcap,下载路径为%s' % (path))
    return response


@csrf_exempt
def pcap_delete(request):
    """
    :describe:  pcap删除
    :param:     http请求
    :return:    响应状态: 成功 未命中(不包含此规则pcap)
    """
    sid = request.POST.get('sid')

    if sid is not None:
        set_del_id(sid)

    rules = RulePcap.objects.all()

    if len(rules) != 0:
        for rule in rules:
            if rule.sid == get_del_id():
                set_pcap(rule.pcap)
                break

    if get_pcap() == "":
        return HttpResponse('该规则无命中的pcap')

    cmd = 'rm -rf ./data/pcaps/' + get_pcap()
    os.system(cmd)
    RulePcap.objects.filter(sid=get_del_id()).delete()

    return HttpResponse('删除成功')


@csrf_exempt
def translate_show(request):
    """
    :describe:  获取要翻译的规则ID
    :param:     http请求
    :return:    输入翻译内容界面
    """
    if request.method == 'POST':
        set_trans_sid(request.POST.get('sid'))
    return render_to_response('translate_rule.html')


def has_chinese(pattern):
    """
    :describe:  判断是否包含中文
    :param:     待判断字符串
    :return:    True: 包含 False: 不包含
    """
    set_default_env()
    for ch in pattern.decode('utf-8'):
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


def random_sid():
    """
    :describe:  随机生成规则ID值(200万-300万)
    :param:     无
    :return:    生成的ID
    """
    sid = random.randint(2e6, 3e6)
    return sid


def encryption(clear_passwd, salt='snort'):
    """
    :describe:      密码加密(sha256算法)
    :param param1:  明文密码
    :param param2:  加密盐
    :return:        加密后密码
    """
    hash_obj = hashlib.sha256()
    clear_passwd += salt
    hash_obj.update(clear_passwd.encode())
    return hash_obj.hexdigest()


@csrf_exempt
def delete(request):
    """
    :describe:  日志删除
    :param:     http请求
    :return:    响应值 0:失败 1:成功
    """
    if request.method == 'POST':
        sid = request.POST.get('id')
        Log.objects.get(id=sid).delete()
        return HttpResponse(1)

    return HttpResponse(0)


def grant_auth():
    """
    :describe:  赋予可执行权限(规则检测工具、系统启动脚本)
    :param:     无
    :return:    无
    """
    cmd = 'chmod +x ./tool/*'
    os.system(cmd)


def set_so_env():
    """
    :describe:  设置规则检测工具依赖库环境变量
    :param:     无
    :return:    无
    """
    cmd = 'source ./tool/config.sh'
    os.system(cmd)


def rules_verify(rule):
    """
    :describe:  验证规则是否被支持
    :param:     待验证规则文件
    :return:    True:支持 False:不支持
    """
    grant_auth()
    set_so_env()
    cmd = './tool/test_tool ' + rule
    result = commands.getstatusoutput(cmd)
    # result = re.findall(r'end\\n(.+?)\'', str(result))
    result = re.findall(r'Rule(.+?):', str(result))
    if len(result) != 0:
        if 'unsupported' in result[0]:
            return False
        else:
            return True
    else:
        return True


def is_hit(rule, pcap, user, ip):
    """
    :describe:      检测规则是否命中pcap
    :param param1:  待检测规则文件
    :param param2:  待检测pcap文件
    :param param3:  当前登录用户 
    :return:        True:命中 False:未命中
    """
    grant_auth()
    set_so_env()
    cmd = './tool/test_tool ' + rule + ' ' + pcap
    result = commands.getstatusoutput(cmd)

    set_default_env()
    if 'hit rules' in str(result):
        record_log(filter(str.isdigit, rule), '规则检测',
                   user, '成功', ip, '%s成功命中规则%s' % (pcap, rule))
        return True
    record_log(filter(str.isdigit, rule), '规则检测',
               user, '失败', ip, '%s未命中规则%s' % (pcap, rule))
    return False


@csrf_exempt
def get_xlsx(request):
    """
    :describe:  导出xlsx表格文件
    :param:     http请求
    :return:    表格文件
    """
    xlsx_path = get_xlsx_path()

    # 向表格写入数据
    write_data(xlsx_path)
    file = None
    try:
        file = open(xlsx_path, 'rb')
    except Exception as e:
        print e
        return HttpResponse('xlsx文件打开失败')

    response = StreamingHttpResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="rules.xlsx"'
    user = request.session['user_name']
    ip = get_ip(request)
    set_default_env()
    record_log('无', '表格导出', user, '成功', ip,
               '导出文件为rules.xlsx,下载路径为%s' % (xlsx_path))

    return response


@csrf_exempt
def time_export(request):
    """
    :describe:  时间范围规则导出
    :param:     http请求
    :return:    导出的文件对象
    """
    rule_list = []
    if request.method == "POST":
        set_date_range(str(request.POST.get('date_range')))
        return HttpResponse(1)

    if request.method == "GET":
        start, end = get_start_end(get_date_range())
        rules_summary = Rule.objects.filter(
            update_time__gte=start, update_time__lte=end)
        for rule in rules_summary:
            rule_obj = CompleteRule.objects.get(sid=rule.sid)
            rule_list.append(rule_obj.rule)
        write_time_export_rules(start, end, rule_list)
    return get_response()


def export_time_range(pwd, rules_path, names_path):
    """
    :describe:      时间范围规则文件下载到本地
    :param param1:  存储被导出的规则和特征文件目录
    :param param2:  被导出规则文件全路径
    :param param3:  被导出特征文件全路径
    :return:        无
    """
    file_list = []
    file_list.append(rules_path)
    file_list.append(names_path)
    path = pwd + 'rules.zip'
    f = zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED)

    for file in file_list:
        f.write(file)
    f.close()
    file = None
    try:
        file = open(path, 'rb')
    except Exception as e:
        print 'error:', e
        return HttpResponse('待导出文件打开出错')
    response = StreamingHttpResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="rules.zip"'
    set_response(response)
    remove_zip_cmd = 'rm -rf ' + path
    os.system(remove_zip_cmd)


def write_time_export_rules(start, end, rule_list):
    """
    :describe:      时间范围内规则写入文件
    :param param1:  开始日期
    :param param2:  结束日期
    :param param3:  在日期范围内规则列表
    :return:        无
    """
    start = str(start).replace(" ", '_')
    end = str(end).replace(" ", '_')
    pwd = get_download_path() + start + end

    if not os.path.exists(pwd):
        os.mkdir(pwd)

    rules_path = pwd + '/export.rules'
    names_path = pwd + '/export'

    for rule in rule_list:
        try:
            with open(rules_path, 'a+') as f:
                f.write(rule + '\n')
        except Exception as e:
            print 'error:', e

    extract_path = get_extract_path()
    cmd = 'python ' + extract_path + ' --rule ' + pwd + ' --out ' + names_path
    os.system(cmd)
    # 压缩文件并下载
    export_time_range(pwd, rules_path, names_path)
    remove_dir_cmd = 'rm -rf ' + pwd
    os.system(remove_dir_cmd)


def write_data(path):
    """
    :describe:  写xlsx表格文件
    :param:     xlsx表格文件存储路径
    :return:    无
    """
    col_names = ['Sid', 'Msg1', 'Msg_ch', 'Msg2', '是否变种', '漏洞利用工具包', '威胁描述', '威胁分类', '恶意代码',
                 '受害软件类型', '受害软件运行环境', '军火工具版本', '病毒名', '查找关键字', 'AVL引擎输出病毒名', '攻击者',
                 '受害者', '控制源', '受控主机', '确认被控', '攻击成功', 'Classtype', 'Rev', 'Reference', 'line']
    wb = Workbook()
    ws = wb.active
    for col in range(0, len(col_names)):
        ws.cell(row=1, column=col + 1, value=col_names[col])

    rule_obj = Rule.objects.filter(contain='是', is_translate='是')

    row = 2
    for rule in rule_obj:
        complete_rule = CompleteRule.objects.get(sid=rule.sid)
        ws.cell(row=row, column=1, value=rule.sid)                  # Sid
        ws.cell(row=row, column=3, value=rule.msg)                  # Msg_ch
        ws.cell(row=row, column=15, value=rule.malname)             # 病毒名
        ws.cell(row=row, column=16, value=rule.attacker)            # 攻击者
        ws.cell(row=row, column=17, value=rule.victim)              # 受害者
        ws.cell(row=row, column=18, value=rule.controller)          # 控制源
        ws.cell(row=row, column=20, value=rule.confirm_controlled)  # 确认被控
        ws.cell(row=row, column=21, value=rule.success_attack)      # 攻击成功
        ws.cell(row=row, column=22, value=rule.class_type)          # Classtype
        ws.cell(row=row, column=23, value=rule.rev)                 # Rev
        ws.cell(row=row, column=24, value=rule.reference)           # Reference
        ws.cell(row=row, column=25, value=complete_rule.rule)       # 完成规则

        row += 1

    wb.save(path)
