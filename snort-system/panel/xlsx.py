# -*- coding: UTF-8 -*-
from openpyxl import Workbook
from models import *


def write_data(path):
    """
    :describe:  写xlsx表格文件
    :param:     xlsx表格文件存储路径
    :return:    无
    """
    col_names = ['Sid', 'Msg1', 'Msg_ch', 'Msg2', '是否变种', '漏洞利用工具包', '威胁描述', '威胁分类', '恶意代码',
                 '受害软件类型', '受害软件运行环境', '军火工具版本', '病毒名', '查找关键字', 'AVL引擎输出病毒名', '攻击者',
                 '受害者', '控制源', '受控主机', '确认被控', '攻击成功', 'Classtype', 'Rev', 'Reference', 'line']
    wb = Workbook()
    ws = wb.active
    for col in range(0, len(col_names)):
        ws.cell(row=1, column=col + 1, value=col_names[col])

    rule_obj = Rule.objects.filter(contain='是', is_translate='是')

    row = 2
    for rule in rule_obj:
        complete_rule = CompleteRule.objects.get(sid=rule.sid)
        ws.cell(row=row, column=1, value=rule.sid)  				# Sid
        ws.cell(row=row, column=3, value=rule.msg)  				# Msg_ch
        ws.cell(row=row, column=15, value=rule.malname)  			# 病毒名
        ws.cell(row=row, column=16, value=rule.attacker) 			# 攻击者
        ws.cell(row=row, column=17, value=rule.victim)  			# 受害者
        ws.cell(row=row, column=18, value=rule.controller)  		# 控制源
        ws.cell(row=row, column=20, value=rule.confirm_controlled)  # 确认被控
        ws.cell(row=row, column=21, value=rule.success_attack)  	# 攻击成功
        ws.cell(row=row, column=22, value=rule.class_type)  		# Classtype
        ws.cell(row=row, column=23, value=rule.rev)  				# Rev
        ws.cell(row=row, column=24, value=rule.reference)  			# Reference
        ws.cell(row=row, column=25, value=complete_rule.rule)  		# 完成规则

        row += 1

    wb.save(path)
